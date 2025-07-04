#!/usr/bin/python
# -*- coding: utf-8 -*-

from __future__ import print_function

from pyVim.connect import SmartConnect, Disconnect
from pyVmomi import vim

import atexit
import ssl
import datetime
from datetime import timedelta,date
import argparse
import socket
from tqdm import tqdm				# Progress bar
from openpyxl import Workbook			# Excel export
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

def main(o):
# accept any certificate here
    context=0
    if hasattr(ssl, 'SSLContext'):
<target>
        context = ssl.SSLContext(ssl.PROTOCOL_TLSv1)
</target>
        context.verify_mode = ssl.CERT_NONE
    try:
        if context:
            si = SmartConnect(host=o["esxserver"],user=o["username"],pwd=o["password"],port=443,sslContext=context)
        else:
            si = SmartConnect(host=o["esxserver"],user=o["username"],pwd=o["password"],port=443)
    except socket.gaierror as e:
        print("cannot connect to "+o["esxserver"])
        return -3
    except Exception as e:
        print("connection error")
        print(type(e))
        return -2
        
    if not si:
        print("Could not connect to the specified host using specified "
              "username and password")
        return -1

    atexit.register(Disconnect, si)

    content = si.RetrieveContent()

    perf_dict = {}
    perfList = content.perfManager.perfCounter
    for counter in perfList:
        counter_full = "{}.{}.{}".format(counter.groupInfo.key, counter.nameInfo.key, counter.rollupType)
        perf_dict[counter_full] = counter.key

#    print(perf_dict)

    vchtime = si.CurrentTime()
#prepare the report date range
    now = datetime.datetime.now()
    
#    (start,end)=getperiod(o)
    end = vchtime - timedelta(minutes=1)
    start = vchtime - timedelta(minutes=(60*24*31))

    report_prefix=str(start.strftime("%Y-%m"))
    
#    print(start,end)

#Gather Host,vm and datastore Data
    object_view = content.viewManager.CreateContainerView(content.rootFolder,[vim.VirtualMachine,vim.HostSystem,vim.Datastore], True)

#open the report files    
    wbvm = {}
    wbvm['wb'] = Workbook()
    wbvm['ws'] = {}
    wbvm['rows'] = {}

    wbhost = {}
    wbhost['wb'] = Workbook()
    wbhost['ws'] = {}
    wbhost['rows'] = {}

    fdatastore = open(report_prefix+"-datastore-report.csv","w")

#collect the data    
    total=len(object_view.view)
    akt=0
    
    if(o["verbose"]):
        pbar=tqdm(total=len(object_view.view))
    
    for obj in object_view.view:
        if(o["verbose"]):
            pbar.update()
        if isinstance(obj,vim.VirtualMachine) or isinstance(obj,vim.HostSystem):
            content = si.RetrieveContent()
            search_index = content.searchIndex
            perfManager = content.perfManager

            metricId = vim.PerformanceManager.MetricId(counterId=perf_dict['cpu.usagemhz.average'], instance="")

            query = vim.PerformanceManager.QuerySpec(entity=obj,
                                                 metricId=[metricId],
                                                 startTime=start, endTime=end,intervalId=7200)
            res=perfManager.QueryPerf(querySpec=[query])

        if isinstance(obj, vim.VirtualMachine):
            if obj.runtime.host.parent.name+" - CPU" not in wbvm['ws']:
                wbvm['ws'][obj.runtime.host.parent.name+" - CPU"] = wbvm['wb'].create_sheet(obj.runtime.host.parent.name+" - CPU")
                wbvm['rows'][obj.runtime.host.parent.name+" - CPU"] = 1
            printVM(obj,res,start,end,wbvm,' - CPU')
            metricId = vim.PerformanceManager.MetricId(counterId=perf_dict['mem.usage.average'], instance="")
            query = vim.PerformanceManager.QuerySpec(entity=obj,
                                                 metricId=[metricId],
                                                 startTime=start, endTime=end,intervalId=7200)
            res=perfManager.QueryPerf(querySpec=[query])
            if obj.runtime.host.parent.name+" - RAM" not in wbvm['ws']:
                wbvm['ws'][obj.runtime.host.parent.name+" - RAM"] = wbvm['wb'].create_sheet(obj.runtime.host.parent.name+" - RAM")
                wbvm['rows'][obj.runtime.host.parent.name+" - RAM"] = 1
            printVM(obj,res,start,end,wbvm," - RAM")
        if isinstance(obj, vim.HostSystem):
            if "ESXHOSTS" not in wbhost['ws']:
                wbhost['ws']["ESXHOSTS"] = wbhost['wb'].create_sheet("ESXHOSTS")
                wbhost['rows']["ESXHOSTS"] = 1
            printHost(obj,res,start,end,wbhost)
        if isinstance(obj, vim.Datastore):
            if(obj.info.vmfs.local==False):                             # Skip all local datastores
                fdatastore.write("{}, {}, {}, {}\n".format(obj.summary.name,
                                             sizeof_fmt(obj.summary.capacity-obj.summary.freeSpace),
                                             sizeof_fmt(obj.summary.freeSpace),
                                             sizeof_fmt(obj.summary.capacity)))         
            
    if(o["verbose"]):
        pbar.close()

    ws = wbvm['wb'].get_sheet_by_name('Sheet')
    if ws is not None:
        wbvm['wb'].remove_sheet(ws)

    ws = wbhost['wb'].get_sheet_by_name('Sheet')
    if ws is not None:
        wbhost['wb'].remove_sheet(ws)

    wbvm['wb'].save(report_prefix+"-CPU-RAM-VM.xls")
    wbhost['wb'].save(report_prefix+"-CPU-Host.xls")
    fdatastore.close()
    
def sizeof_fmt(num):
    return "%3.1f" % (num/(1024*1024*1024.0))
    
def printVM(obj,res,start,end,wb,stat):
    dates = []
    values = []

    if res and res[0].sampleInfo:
        if wb['rows'][obj.runtime.host.parent.name+stat]== 1:
            for d in daterange(start):
                dates.append(str(d.date()))
            row=["VM"]+dates+["","accumulated"]
            wb['ws'][obj.runtime.host.parent.name+stat].append(row)
            wb['rows'][obj.runtime.host.parent.name+stat] +=1

        dates=[]
        for d in range(0,len(res[0].sampleInfo)-1):
            dates.append(res[0].sampleInfo[d].timestamp.replace(tzinfo=None))
        
        accumulated=0
        i=0
        for d in daterange(start):
            if d not in dates:
                values.append(0)
            else:
                values.append(float(res[0].value[0].value[i])/100)
                accumulated += float(res[0].value[0].value[i])/100
                i+=1
        row=[obj.name]+values+[""]+[accumulated]
        wb['ws'][obj.runtime.host.parent.name+stat].append(row)

    return

def printHost(obj,res,start,end,wb):
    dates = []
    values = []

    if res and res[0].sampleInfo:
        if wb['rows']['ESXHOSTS']== 1:
            for d in daterange(start):
                dates.append(str(d))
            row=["Host"]+dates+["","accumulated"]
            wb['ws']['ESXHOSTS'].append(row)
            wb['rows']['ESXHOSTS'] +=1

        dates=[]
        for d in range(0,len(res[0].sampleInfo)-1):
            dates.append(res[0].sampleInfo[d].timestamp.replace(tzinfo=None))
        
        accumulated=0
        i=0
        for d in daterange(start):
            if d not in dates:
                values.append(0)
            else:
                values.append(float(res[0].value[0].value[i])/100)
                accumulated += float(res[0].value[0].value[i])/100
                i+=1
        row=[obj.name]+values+[""]+[accumulated]
        wb['ws']['ESXHOSTS'].append(row)

    return

#iterator for date ranges    
def daterange(start_date):
   now = datetime.datetime.now()
   start_date=datetime.datetime(now.year,now.month,now.day,0,0,0)
   start_date=start_date-timedelta(days=31)

   for n in range(0,(31*24/2)):	# every 2 hours
      yield start_date + timedelta(minutes=n*120)


		 
#simple config parser         
def parseConfig(filename):
   COMMENT_CHAR = '#'
   OPTION_CHAR =  '='
   options = {}
   f = open(filename)
   for line in f:
      if COMMENT_CHAR in line:
         line, comment = line.split(COMMENT_CHAR, 1)
      if OPTION_CHAR in line:
         option, value = line.split(OPTION_CHAR, 1)
         option = option.strip()
         value = value.strip()
         options[option] = value
   f.close()

   return options

def getperiod(o):
    now = datetime.datetime.now()
    if o["year"]==None:
        year=now.year
    else:
        year=int(o["year"])
    month=None
    if o["month"]!=None:
        o["month"]=o["month"][:3]
        if o["month"].upper()=="JAN":
            month=1
        elif o["month"].upper()=="FEB":
            month=2
        elif o["month"].upper()=="MAR" or o["month"].upper()=="MÄR":
            month=3
        elif o["month"].upper()=="APR":
            month=4        
        elif o["month"].upper()=="MAY" or o["month"].upper()=="MAI":
            month=5
        elif o["month"].upper()=="JUN":
            month=6
        elif o["month"].upper()=="JUL":
            month=7
        elif o["month"].upper()=="AUG":
            month=8
        elif o["month"].upper()=="SEP":
            month=9
        elif o["month"].upper()=="OCT" or o["month"].upper()=="OKT":
            month=10
        elif o["month"].upper()=="NOV":
            month=11
        elif o["month"].upper()=="DEC" or o["month"].upper()=="DEZ":
            month=12
    else:
        now = datetime.datetime.now()
        month=now.month-1
        
    if(month==0):
        month=12
        year-=1
        
    start=datetime.datetime(year,month,1,0,0)
    
    month+=1
    
    if(month==13):
        month=1
        year+=1
        
    end=datetime.datetime(year,month,1,0,0)

    return(start,end)
   
# Start program
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-c","--config", help="configfile")
    parser.add_argument("-v","--verbose", help="show statusbar",action="store_true")
    args = parser.parse_args()
    
    if(args.config==None):
        config="config.ini"
    else:
        config=args.config
    
    c=parseConfig(config)
#    c["month"]=args.month
#    c["year"]=args.year
    c["verbose"]=args.verbose
    main(c)
